import openpyxl

# ① 엑셀 만들기
wb = openpyxl.Workbook()

# ② 엑셀 워크시트 만들기
ws = wb.create_sheet('SquidGame')

# ③ 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'
ws['A2'] = 1
ws['B2'] = '오일남'

# ④ 엑셀 저장하기
wb.save('/Users/kyle/Documents/Study/Python/StartCoding/workspace/excel/entry_data.xlsx')
