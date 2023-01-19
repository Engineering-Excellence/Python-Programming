import openpyxl

filePath = '/Users/kyle/Documents/Study/Python/StartCoding/workspace/excel/entry_data.xlsx'

# ① 엑셀 불러오기
wb = openpyxl.load_workbook(filePath)

# ② 엑셀 시트 선택
ws = wb['SquidGame']

# ③ 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# ④ 엑셀 저장하기
wb.save(filePath)
