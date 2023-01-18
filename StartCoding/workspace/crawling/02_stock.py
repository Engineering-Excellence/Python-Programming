import requests
from bs4 import BeautifulSoup
import openpyxl

filePath = '/Users/kyle/Documents/Study/Python/StartCoding/workspace/crawling/stock.xlsx'

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filePath)
# ws = wb.create_sheet('stock')
ws = wb['stock']

ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'

codes = [
    '005930',
    '000660',
    '035720'
]

for i in range(len(codes)):
    url = f'https://finance.naver.com/item/sise.naver?code={codes[i]}'
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    title = soup.select_one('div.wrap_company > h2 > a').text
    price = int(soup.select_one('#_nowVal').text.replace(',', ''))
    bid_price = int(soup.select_one('span.tah.p11').text.replace(',', ''))
    print(f'종목: {title}, 현재가: {price}, 매수호가: {bid_price}')
    ws[f'A{i + 2}'] = title
    ws[f'B{i + 2}'] = price

# wb.save(filePath)
